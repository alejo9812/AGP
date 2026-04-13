from __future__ import annotations

import csv
import io
import uuid
from collections.abc import Iterable
from datetime import datetime, timedelta
from decimal import Decimal

from fastapi import HTTPException, UploadFile, status
from openpyxl import load_workbook
from sqlalchemy import delete, select
from sqlalchemy.orm import Session

from app.core.config import get_settings
from app.models import (
    AuditEventType,
    AuditLog,
    Customer,
    InventoryImportBatch,
    InventoryItem,
    OperationalStatus,
    ProductType,
    QrTag,
    SourceSetStatus,
    StatusCatalog,
    User,
    UserRole,
    Vehicle,
    WarehouseLocation,
)
from app.services.data_quality import build_review_reasons, summarize_duplicates

REQUIRED_HEADERS = [
    "ID",
    "OrderID",
    "Serial",
    "Vehicle",
    "Created",
    "Product",
    "Invoice",
    "InvoiceCost",
    "Customer",
    "DaysStored",
    "SetStatus",
]


def excel_serial_to_date(value: str | int | float | None) -> datetime | None:
    if value in (None, ""):
        return None
    try:
        serial = int(float(str(value)))
    except ValueError:
        try:
            return datetime.fromisoformat(str(value))
        except ValueError:
            return None
    return datetime(1899, 12, 30) + timedelta(days=serial)


def normalize_header(value: str | None) -> str:
    return (value or "").strip()


def detect_source_type(filename: str) -> str:
    lowered = filename.lower()
    if lowered.endswith(".xlsx"):
        return "xlsx"
    if lowered.endswith(".csv"):
        return "csv"
    raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Solo se soportan archivos CSV y XLSX.")


def parse_rows(file_name: str, payload: bytes) -> list[dict[str, str | None]]:
    source_type = detect_source_type(file_name)
    if source_type == "csv":
        content = payload.decode("utf-8-sig")
        reader = csv.DictReader(io.StringIO(content))
        return [{key: value for key, value in row.items()} for row in reader]

    workbook = load_workbook(filename=io.BytesIO(payload), read_only=True, data_only=True)
    worksheet = workbook[workbook.sheetnames[0]]
    rows = list(worksheet.iter_rows(values_only=True))
    headers = [normalize_header(str(value)) for value in rows[0]]
    data: list[dict[str, str | None]] = []
    for raw_row in rows[1:]:
        data.append({header: None if value is None else str(value) for header, value in zip(headers, raw_row)})
    return data


def ensure_master_data(db: Session) -> None:
    settings = get_settings()

    if db.scalar(select(WarehouseLocation).where(WarehouseLocation.code == settings.default_warehouse_location)) is None:
        db.add(WarehouseLocation(code=settings.default_warehouse_location, zone="MX", aisle="UNASSIGNED", level="0"))

    status_catalog = {
        OperationalStatus.IN_STOCK: "Disponible en bodega.",
        OperationalStatus.RESERVED: "Reservado temporalmente para agrupacion o despacho.",
        OperationalStatus.GROUPED: "Ya fue utilizado como donante.",
        OperationalStatus.COMPLETED: "Set completado y listo para flujo posterior.",
        OperationalStatus.READY_FOR_DISPATCH: "Listo para despacho.",
        OperationalStatus.DISPATCHED: "Despachado.",
        OperationalStatus.BLOCKED: "Bloqueado por inconsistencia o accion operativa.",
        OperationalStatus.REVIEW_NEEDED: "Requiere revision manual.",
    }
    for status_code, description in status_catalog.items():
        if db.scalar(select(StatusCatalog).where(StatusCatalog.status_code == status_code)) is None:
            db.add(StatusCatalog(status_code=status_code, display_name=status_code.value, description=description))

    demo_users = [
        ("b111a0c8-1f02-4979-b65f-734314bf93b6", "Daniela Vargas", "daniela.vargas@agp.demo", UserRole.ADMIN),
        ("d7bd4e61-e253-47bc-bf26-7bd482cc8388", "Luis Herrera", "luis.herrera@agp.demo", UserRole.COMMERCIAL_ANALYST),
        ("19383d57-e3b0-4f3d-b17b-98935c81232d", "Paola Garcia", "paola.garcia@agp.demo", UserRole.WAREHOUSE_OPERATOR),
        ("b2209b20-1651-4257-b242-15ecce661dce", "Sofia Martinez", "sofia.martinez@agp.demo", UserRole.EXECUTIVE_READONLY),
    ]
    for user_uuid, full_name, email, role in demo_users:
        if db.scalar(select(User).where(User.email == email)) is None:
            db.add(User(user_uuid=user_uuid, full_name=full_name, email=email, role=role))
    db.commit()


def preview_dataset(file_name: str, payload: bytes) -> dict:
    rows = parse_rows(file_name, payload)
    detected_headers = list(rows[0].keys()) if rows else []
    return {
        "source_type": detect_source_type(file_name),
        "detected_headers": detected_headers,
        "missing_required_headers": [header for header in REQUIRED_HEADERS if header not in detected_headers],
        "preview_rows": rows[:5],
        "row_count": len(rows),
    }


def wipe_inventory_state(db: Session) -> None:
    for model in (AuditLog, QrTag, InventoryItem, InventoryImportBatch):
        db.execute(delete(model))
    db.commit()


def get_or_create_customer(db: Session, name: str | None) -> Customer | None:
    if not name:
        return None
    record = db.scalar(select(Customer).where(Customer.name == name))
    if record:
        return record
    record = Customer(name=name)
    db.add(record)
    db.flush()
    return record


def get_or_create_vehicle(db: Session, name: str | None) -> Vehicle | None:
    if not name:
        return None
    record = db.scalar(select(Vehicle).where(Vehicle.name == name))
    if record:
        return record
    record = Vehicle(name=name)
    db.add(record)
    db.flush()
    return record


def get_or_create_product(db: Session, name: str) -> ProductType:
    record = db.scalar(select(ProductType).where(ProductType.name == name))
    if record:
        return record
    record = ProductType(name=name)
    db.add(record)
    db.flush()
    return record


def default_location(db: Session) -> WarehouseLocation:
    settings = get_settings()
    location = db.scalar(select(WarehouseLocation).where(WarehouseLocation.code == settings.default_warehouse_location))
    if location is None:
        location = WarehouseLocation(code=settings.default_warehouse_location)
        db.add(location)
        db.flush()
    return location


def normalize_rows(rows: Iterable[dict[str, str | None]]) -> list[dict[str, str | int | float | None]]:
    normalized: list[dict[str, str | int | float | None]] = []
    for row in rows:
        normalized.append(
            {
                "ID": row.get("ID"),
                "OrderID": row.get("OrderID"),
                "Serial": row.get("Serial"),
                "Vehicle": normalize_header(row.get("Vehicle")),
                "Created": row.get("Created"),
                "Product": normalize_header(row.get("Product")),
                "Invoice": row.get("Invoice"),
                "InvoiceCost": float(row.get("InvoiceCost") or 0),
                "Customer": normalize_header(row.get("Customer")),
                "DaysStored": int(float(row.get("DaysStored") or 0)),
                "SetStatus": normalize_header(row.get("SetStatus")),
            }
        )
    return normalized


def process_import(
    db: Session,
    actor: User,
    upload: UploadFile,
    replace_existing: bool = True,
) -> dict:
    payload = upload.file.read()
    rows = parse_rows(upload.filename or "inventory.xlsx", payload)
    detected_headers = list(rows[0].keys()) if rows else []
    missing_headers = [header for header in REQUIRED_HEADERS if header not in detected_headers]
    if missing_headers:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=f"Faltan columnas requeridas: {', '.join(missing_headers)}",
        )

    ensure_master_data(db)
    if replace_existing:
        wipe_inventory_state(db)
        ensure_master_data(db)

    normalized_rows = normalize_rows(rows)
    duplicate_stats = summarize_duplicates(normalized_rows)
    batch = InventoryImportBatch(
        batch_uuid=str(uuid.uuid4()),
        file_name=upload.filename or "inventory.xlsx",
        source_type=detect_source_type(upload.filename or "inventory.xlsx"),
        total_rows=len(normalized_rows),
        status="processed",
    )
    db.add(batch)
    db.flush()

    location = default_location(db)
    review_items = 0
    free_stock_items = 0

    for row in normalized_rows:
        customer_name = str(row.get("Customer") or "").strip() or None
        vehicle_name = str(row.get("Vehicle") or "").strip() or None
        product_name = str(row.get("Product") or "").strip()
        review_reasons = build_review_reasons(row)
        needs_review = any(
            "Vehicle vacio" in reason or "Product vacio" in reason or "SetStatus invalido" in reason
            for reason in review_reasons
        )
        is_free_stock = customer_name is None

        if needs_review:
            review_items += 1
        if is_free_stock:
            free_stock_items += 1

        customer = get_or_create_customer(db, customer_name)
        vehicle = get_or_create_vehicle(db, vehicle_name)
        product = get_or_create_product(db, product_name or "UNKNOWN")
        created_date = excel_serial_to_date(row.get("Created"))

        item = InventoryItem(
            source_id=str(row.get("ID")),
            order_id=str(row.get("OrderID")),
            serial=str(row.get("Serial")),
            vehicle_name=vehicle_name,
            created_date=created_date.date() if created_date else None,
            product_name=product_name or "UNKNOWN",
            invoice=str(row.get("Invoice")),
            invoice_cost=Decimal(str(row.get("InvoiceCost") or 0)),
            customer_name=customer_name,
            days_stored=int(row.get("DaysStored") or 0),
            set_status=SourceSetStatus(str(row.get("SetStatus"))),
            operational_status=OperationalStatus.REVIEW_NEEDED if needs_review else OperationalStatus.IN_STOCK,
            is_free_stock=is_free_stock,
            needs_review=needs_review,
            review_reasons=review_reasons,
            customer_id=customer.id if customer else None,
            vehicle_id=vehicle.id if vehicle else None,
            product_type_id=product.id,
            import_batch_id=batch.id,
            location_id=location.id,
        )
        db.add(item)
        db.flush()
        db.add(QrTag(qr_token=f"AGP-{item.source_id[-8:]}".upper(), inventory_item_id=item.id))

    batch.valid_rows = len(normalized_rows) - review_items
    batch.rows_needing_review = review_items
    db.add(
        AuditLog(
            event_uuid=str(uuid.uuid4()),
            event_type=AuditEventType.IMPORT,
            actor_user_id=actor.id,
            details={
                "batch_uuid": batch.batch_uuid,
                "total_rows": len(normalized_rows),
                "duplicate_orders": duplicate_stats["duplicate_orders"],
                "duplicate_serials": duplicate_stats["duplicate_serials"],
            },
        )
    )
    db.commit()
    db.refresh(batch)

    return {
        "batch": batch,
        "inserted_items": len(normalized_rows),
        "review_items": review_items,
        "free_stock_items": free_stock_items,
        "message": "Importacion completada y lista para analisis.",
        "duplicate_stats": duplicate_stats,
    }
