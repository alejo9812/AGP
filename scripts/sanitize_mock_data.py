from __future__ import annotations

import csv
import hashlib
import json
import sys
from pathlib import Path

from openpyxl import load_workbook


CUSTOMER_CACHE: dict[str, str] = {}
VEHICLE_CACHE: dict[str, str] = {}


def anonymize(value: str, prefix: str, cache: dict[str, str]) -> str:
    if not value:
        return ""
    if value not in cache:
        cache[value] = f"{prefix} {len(cache) + 1:02d}"
    return cache[value]


def token(value: str, prefix: str) -> str:
    digest = hashlib.sha1(value.encode("utf-8")).hexdigest()[:10].upper()
    return f"{prefix}{digest}"


def main(source_path: Path, output_dir: Path) -> None:
    workbook = load_workbook(source_path, read_only=True, data_only=True)
    worksheet = workbook[workbook.sheetnames[0]]
    rows = list(worksheet.iter_rows(values_only=True))
    headers = [str(value) for value in rows[0]]
    sample_rows = []
    counters = {"total_rows": 0, "free_stock": 0, "needs_review": 0}

    for raw in rows[1:241]:
        record = dict(zip(headers, raw))
        counters["total_rows"] += 1
        customer = str(record.get("Customer") or "")
        vehicle = str(record.get("Vehicle") or "")
        if not customer:
            counters["free_stock"] += 1
        if not vehicle:
            counters["needs_review"] += 1
        sample_rows.append(
            {
                "ID": token(str(record["ID"]), "ID-"),
                "OrderID": token(str(record["OrderID"]), "ORD-"),
                "Serial": token(str(record["Serial"]), "SER-"),
                "Vehicle": anonymize(vehicle, "Vehicle", VEHICLE_CACHE),
                "Created": str(record["Created"]),
                "Product": record["Product"],
                "Invoice": token(str(record["Invoice"]), "INV-"),
                "InvoiceCost": record["InvoiceCost"],
                "Customer": anonymize(customer, "Customer", CUSTOMER_CACHE),
                "DaysStored": record["DaysStored"],
                "SetStatus": record["SetStatus"],
            }
        )

    output_dir.mkdir(parents=True, exist_ok=True)
    csv_path = output_dir / "agp_inventory_sample.csv"
    json_path = output_dir / "agp_inventory_sample.json"

    with csv_path.open("w", encoding="utf-8", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=headers)
        writer.writeheader()
        writer.writerows(sample_rows)

    json_payload = {
        "summary": counters,
        "rows": sample_rows[:20],
    }
    json_path.write_text(json.dumps(json_payload, indent=2), encoding="utf-8")
    print(f"Wrote {csv_path}")
    print(f"Wrote {json_path}")


if __name__ == "__main__":
    source = Path(sys.argv[1]) if len(sys.argv) > 1 else Path(r"C:\Users\hmont\Downloads\Mock_Data.xlsx")
    destination = Path(sys.argv[2]) if len(sys.argv) > 2 else Path("data/sample")
    main(source, destination)
